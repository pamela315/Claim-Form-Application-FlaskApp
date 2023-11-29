from flask import Flask
from flask import request
from flask import render_template
from googleads import ad_manager
from googleads import errors
from datetime import date, timedelta
from datetime import datetime
import pandas as pd
import _locale
import gspread
from gspread_formatting  import *
from gspread_dataframe import get_as_dataframe, set_with_dataframe
import os
import glob
import openpyxl
from openpyxl import load_workbook, Workbook
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import numpy as np

app = Flask(__name__)
app.debug = True


############### Google Sheet API Authentication ###################


scope = ['https://spreadsheets.google.com/feeds',
     'https://www.googleapis.com/auth/drive']

THIS_FOLDER2 = os.path.dirname(os.path.abspath(__file__))
creds2 = os.path.join(THIS_FOLDER2, 'creds2.json')

creds = ServiceAccountCredentials.from_json_keyfile_name(creds2,scope)
client = gspread.authorize(creds)



@app.route('/')
def index():
    return render_template("main_page.html")


@app.route('/result', methods=['post','get'])
def result():
    if request.method == 'POST':
        team = request.form.get('team')
        payto = request.form.get('payto')
        amount = request.form.get('amount')
        adno = request.form.get('adno')
        vendorcode = request.form.get('vendorcode')
        appliedby = request.form.get('appliedby')
        extno = request.form.get('extno')
        option = request.form.get('type')
        payment = request.form.get('method')
        mail = request.form.get('mail')
        otherSpecify = request.form.get('otherSpecify')
        currency = request.form.get('currency')
        urgent = request.form.get('urgent')
        urgentdate = request.form.get('urgentdate')
        service = request.form.get('service')
        serviceOther = request.form.get('ServiceOther')
        itemname = request.form.get('itemname')
        itemOther = request.form.get('itemOther')
        publishedDate = request.form.get('publishedDate')
        clientname = request.form.get('clientname')
        password = request.form.get('password')
        details = request.form.get('details')
        teamsOthers = request.form.get('teamsOthers')


        f_reason = request.form.get('f_reason')
        f_reasonOther = request.form.get('f_reasonOther')
        f_deadline = request.form.get('f_deadline')
        f_chinesename = request.form.get('f_chinesename')
        f_engname = request.form.get('f_engname')
        f_cal = request.form.get('f_cal')
        calOther = request.form.get('calOther')
        f_date = request.form.get('f_date')


############### Excel Write data ###################


    THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
    my_file = os.path.join(THIS_FOLDER, 'templates/AC002_Template.xlsx')
    outsource_file = os.path.join(THIS_FOLDER, 'templates/Outsourced_Template.xlsx')

    now =  (datetime.now() + timedelta(hours=8)).strftime("%Y%m%d%H%M%S")


    book = openpyxl.load_workbook(my_file)
    outsource_book = openpyxl.load_workbook(outsource_file)

    sheet = book.active

    sheet['B14'] = payto
    sheet['B16'] = vendorcode
    amountFinal = "{:,.2f}".format(float(amount))
    sheet['C18'] = currency + str(amountFinal)
    sheet['D34'] = 'This Amount has been deducted from advertising revenue (Ad #' + adno +')'
    sheet['B21'] = clientname


    if team == 'ADV':
        sheet['D8'] = 'ADV'
    elif team == 'HKET_Adv':
        sheet['D8'] = 'HKET Adv'
    elif team == 'iET_A':
        sheet['D8'] = 'iET - A'
    elif team == 'iET_B':
        sheet['D8'] = 'iET - B'
    elif team == 'iET_C':
        sheet['D8'] = 'iET - C'
    elif team == 'iET_D':
        sheet['D8'] = 'iET - D'
    elif team == 'iET_E':
        sheet['D8'] = 'iET - E'
    elif team == 'PET_A':
        sheet['D8'] = 'PET - A'
    elif team == 'PET_B':
        sheet['D8'] = 'PET - B'
    elif team == 'PET_C':
        sheet['D8'] = 'PET - C'
    elif team == 'PET_D':
        sheet['D8'] = 'PET - D'
    elif team == 'PET_E':
        sheet['D8'] = 'PET - E'
    elif team == 'SkyPost_Adv':
        sheet['D8'] = 'Sky Post Adv'
    elif team == 'Sky_Marketing':
        sheet['D8'] = 'Sky Marketing'
    elif team == 'Sky_A':
        sheet['D8'] = 'Sky A'
    elif team == 'Sky_B':
        sheet['D8'] = 'Sky B'
    elif team == 'Sky_C':
        sheet['D8'] = 'Sky C'
    elif team == 'Sky_D':
        sheet['D8'] = 'Sky D'
    elif team == 'Sky_E':
        sheet['D8'] = 'Sky E'
    elif team == 'UL':
        sheet['D8'] = 'UL'
    elif team == 'UM':
        sheet['D8'] = 'UM'
    elif team == 'UT':
        sheet['D8'] = 'UT'
    elif team == 'UFB':
        sheet['D8'] = 'UFB'
    elif team == 'UHK':
        sheet['D8'] = 'UHK'
    elif team == 'Others':
        sheet['D8'] = teamsOthers
        team = teamsOthers
    elif team == 'Ezone':
        sheet['D8'] = 'Ezone'
    elif team == 'iMoney':
        sheet['D8'] = 'iMoney'
    else:
        sheet['D8'] = team



    if service == 'others':
        sheet['B22'] = serviceOther
    else:
        sheet['B22'] = service

    if itemname == 'others':
        sheet['B23'] = itemOther
    else:
        sheet['B23'] = itemname

    sheet['B24'] = publishedDate
    sheet['B25'] = details
    sheet['A34'] = appliedby
    sheet['A35'] = 'Applicant\'s Name (Ext. No.' + extno + ' )'

    if payment == 'cheque':
        sheet['A11'] = 'Payment Method :      ☑ Cheque           ☐ T/T           ☐  Other (Specify)'
    elif payment == 'TT':
        sheet['A11'] = 'Payment Method :      ☐ Cheque           ☑ T/T           ☐  Other (Specify)'
    elif payment == 'other':
        sheet['A11'] = 'Payment Method :      ☐ Cheque           ☐ T/T           ☑  Other (Specify):'+ otherSpecify
    else:
        sheet['A11'] = 'Payment Method :      ☑ Cheque           ☐ T/T           ☐  Other (Specify)'


    if mail == 'fin' and  urgent == 'yes':
        sheet['A12'] ='Mail / Send by :       ☑ FIN                     ☐ Applicant             ☑ Urgent (Date Required) :' + urgentdate
    elif mail == 'fin' and  urgent == 'no':
        sheet['A12'] ='Mail / Send by :       ☑ FIN                     ☐ Applicant             ☐ Urgent (Date Required) :'
    elif mail == 'applicant' and  urgent == 'yes':
        sheet['A12'] = 'Mail / Send by :       ☐ FIN                     ☑ Applicant             ☑ Urgent (Date Required) :' + urgentdate
    elif mail == 'applicant' and  urgent == 'no':
        sheet['A12'] = 'Mail / Send by :       ☐ FIN                     ☑ Applicant             ☐ Urgent (Date Required) :'
    else:
        sheet['A12'] = 'Mail / Send by :       ☐ FIN                     ☑ Applicant             ☐ Urgent (Date Required) '


    medium = Side(border_style="medium", color="000000")
    mediumRed = Side(border_style="medium", color="FF0000")



    sheet['B21'].border = Border(bottom=medium)
    sheet['C21'].border = Border(bottom=medium)
    sheet['D21'].border = Border(bottom=medium)


    sheet['B22'].border = Border(bottom=medium)
    sheet['C22'].border = Border(bottom=medium)
    sheet['D22'].border = Border(bottom=medium)


    sheet['B23'].border = Border(bottom=medium)
    sheet['C23'].border = Border(bottom=medium)
    sheet['D23'].border = Border(bottom=medium)


    sheet['B24'].border = Border(bottom=medium)
    sheet['C24'].border = Border(bottom=medium)
    sheet['D24'].border = Border(bottom=medium)


    sheet['B25'].border = Border(bottom=medium)
    sheet['C25'].border = Border(bottom=medium)
    sheet['D25'].border = Border(bottom=medium)


    sheet['D36'].border = Border(left=mediumRed, right=mediumRed)
    sheet['D37'].border = Border(bottom=mediumRed, left=mediumRed, right=mediumRed)

############### OutSource File ###################

    outsource_sheet = outsource_book.active
    outsource_sheet['H4'] = appliedby

    if f_reason == 'extraworkload':
        outsource_sheet['B7'] = '額外工作'
    elif f_reason == 'AL':
        outsource_sheet['B7'] = '同事放假'
    else:
        outsource_sheet['B7'] = f_reasonOther


    if service == 'others':
        service_final = serviceOther
    else:
        service_final = service

    if itemname == 'others':
        itemname_final = itemOther
    else:
        itemname_final = itemname



    Details_Final = '客戶名稱:' + clientname + ',' + '服務性質:' + service_final + ',' + '項目名稱:' + itemname_final + ',' + '刊登日期:' + publishedDate

    outsource_sheet['B14'] = Details_Final

    outsource_sheet['B17'] = f_deadline
    outsource_sheet['B20'] = currency + str(amountFinal)
    outsource_sheet['C32'] = f_chinesename
    outsource_sheet['H32'] = f_engname
    outsource_sheet['B34'] = f_date

    if f_cal == 'hour':
        outsource_sheet['D35'] = '按時間'
    elif f_cal == 'job':
        outsource_sheet['D35'] = '按工作'
    else:
        outsource_sheet['D35'] = calOther

    outsource_sheet['B37'] = currency + str(amountFinal)


############### Google Sheet Write Data###################


    month = (datetime.now() + timedelta(hours=8)).strftime("%b")
    year = (datetime.now() + timedelta(hours=8)).strftime("%y")
    today = (datetime.now() + timedelta(hours=8)).strftime("%d/%m/%y")

    worksheet_name =  month + ' ' + year


    if option == 'vendor':
        try:
            spreadsheet_V = client.open('AC002 Vendor')
            g_sheet = client.open('AC002 Vendor').worksheet(worksheet_name)
            all = g_sheet.get_all_values()
            endow = len(all) + 1
            if endow <= 10:
                login_no = month + ' ' + year + ' - C0' + str(len(all))
            else:
                login_no = month + ' ' + year + ' - C' + str(len(all))
        except:
            spreadsheet_V.add_worksheet(title=worksheet_name, rows="10000", cols="21")
            g_sheet = client.open('AC002 Vendor').worksheet(worksheet_name)
            first_column = {'Log In No.':[],'Team':[],'Applied Date':[],'Pay To':[],'Currency':[],	'Amount $':[],	'Ad. No.':[],	'Name of Client':[],	'Nature of Service':[],	'Name of Item':[],'Published Date':[],'Details':[],'vendor':[],'Applied By':[],'Diane':[],'Transfer Dept.Date':[],'Cheque Receiver':[],'Changes':[],'Last Update Date':[],'Last Updated by':[],'Enquiry Code':[]}
            df2 = pd.DataFrame(data=first_column)
            set_with_dataframe(g_sheet, df2,row=1, col=1, include_index=False, include_column_header=True,resize=False, allow_formulas=True)
            login_no = month + ' ' + year + ' - C01'
            all = g_sheet.get_all_values()
            endow = len(all) + 1

    elif option == 'freelance':
        try:
            spreadsheet_F = client.open('AC002 Freelance')
            g_sheet = client.open('AC002 Freelance').worksheet(worksheet_name)
            all = g_sheet.get_all_values()
            endow = len(all) + 1
            if endow <= 10:
                login_no = month + ' ' + year + ' - F0' + str(len(all))
            else:
                login_no = month + ' ' + year + ' - F' + str(len(all))
        except:
            spreadsheet_F.add_worksheet(title=worksheet_name, rows="10000", cols="21")
            g_sheet = client.open('AC002 Freelance').worksheet(worksheet_name)
            first_column = {'Log In No.':[],'Team':[],'Applied Date':[],'Pay To':[],'Currency':[],	'Amount $':[],	'Ad. No.':[],	'Name of Client':[],	'Nature of Service':[],	'Name of Item':[],'Published Date':[],'Details':[],'vendor':[],'Applied By':[],'Diane':[],'Transfer Dept.Date':[],'Cheque Receiver':[],'Changes':[],'Last Update Date':[],'Last Updated by':[],'Enquiry Code':[]}
            df2 = pd.DataFrame(data=first_column)
            set_with_dataframe(g_sheet, df2,row=1, col=1, include_index=False, include_column_header=True,resize=False, allow_formulas=True)
            login_no = month + ' ' + year + ' - F01'
            all = g_sheet.get_all_values()
            endow = len(all) + 1
    elif option == 'staff':
        try:
            spreadsheet_S = client.open('AC002 Staff')
            g_sheet = client.open('AC002 Staff').worksheet(worksheet_name)
            all = g_sheet.get_all_values()
            endow = len(all) + 1
            if endow <= 10:
                login_no = month + ' ' + year + ' - S0' + str(len(all))
            else:
                login_no = month + ' ' + year + ' - S' + str(len(all))
        except:
            spreadsheet_S.add_worksheet(title=worksheet_name, rows="10000", cols="21")
            g_sheet = client.open('AC002 Staff').worksheet(worksheet_name)
            first_column = {'Log In No.':[],'Team':[],'Applied Date':[],'Pay To':[],'Currency':[],	'Amount $':[],	'Ad. No.':[],	'Name of Client':[],	'Nature of Service':[],	'Name of Item':[],'Published Date':[],'Details':[],'vendor':[],'Applied By':[],'Diane':[],'Transfer Dept.Date':[],'Cheque Receiver':[],'Log':[],'Last Update Date':[],'Last Updated by':[],'Enquiry Code':[]}
            df2 = pd.DataFrame(data=first_column)
            set_with_dataframe(g_sheet, df2,row=1, col=1, include_index=False, include_column_header=True,resize=False, allow_formulas=True)
            login_no = month + ' ' + year + ' - S01'
            all = g_sheet.get_all_values()
            endow = len(all) + 1
    else:
        spreadsheet_F = client.open('AC002 Freelance')
        g_sheet = client.open('AC002 Freelance').worksheet(worksheet_name)
        all = g_sheet.get_all_values()
        endow = len(all) + 1
        if endow <= 10:
            login_no = month + ' ' + year + ' - F0' + str(len(all))
        else:
            login_no = month + ' ' + year + ' - F' + str(len(all))



    d = {'Log In No.':[login_no],'Team':[team],'Applied Date':[today],'Pay To':[payto],'Currency':[currency],'Amount $':[amount],'Ad. No.':[adno],'Name of Client':[clientname],'Nature of Service':[service_final],'Name of Item':[itemname_final],'Published Date':[publishedDate],'Details':[details],'vendor':[vendorcode],'Applied By':[appliedby],'Diane':[' '],'Transfer Dept.Date':[' '],'Cheque Receiver':[' '],'Log':[' '],'Last Update Date':[' '],'Last Updated by':[' '],'Enquiry Code':[password]}

    df = pd.DataFrame(data=d)
    set_with_dataframe(g_sheet, df,row=endow, col=1, include_index=False, include_column_header=False,
                       resize=False, allow_formulas=True)

############### Save Excel ###################


    sheet['D2'] = 'Login No.: ' + login_no
    sheet['B32'] = today
    outsource_sheet['C5'] = today

    savefile = 'AC002/AC002_' + appliedby + '_' + login_no + '.xlsx'

    link = 'http://pamelatsui.pythonanywhere.com/' + savefile

    AC002_file = os.path.join(THIS_FOLDER, savefile)

    book.save(AC002_file)

    outsource_link = ''
    if option == 'freelance':
        outsource_savefile = 'OutSource/Outsource_' + appliedby + '_' + login_no + '.xlsx'
        outsource_link = 'http://pamelatsui.pythonanywhere.com/' + outsource_savefile
        outsource_file = os.path.join(THIS_FOLDER, outsource_savefile)
        outsource_book.save(outsource_file)


    return render_template("result.html", login_no = login_no, link = link, payment = payment, mail = mail, outsource_link = outsource_link)


@app.route('/edit/')
def edit():
    return render_template("edit.html")


############### Google Sheet Write Data###################

@app.route('/changes', methods=['post','get'])
def changes():
    msg = ''
    revised_link = ''
    revised_outsource_link = ''
    month = (datetime.now() + timedelta(hours=8)).strftime("%b")
    year = (datetime.now() + timedelta(hours=8)).strftime("%y")
    today = (datetime.now() + timedelta(hours=8)).strftime("%d/%m/%y")

    if request.method == 'POST':
        loginno = request.form.get('loginno')
        team = request.form.get('team')
        payto = request.form.get('payto')
        amount = request.form.get('amount')
        adno = request.form.get('adno')
        vendorcode = request.form.get('vendorcode')
        appliedby = request.form.get('appliedby')
        extno = request.form.get('extno')
        option = request.form.get('type')
        payment = request.form.get('method')
        mail = request.form.get('mail')
        otherSpecify = request.form.get('otherSpecify')
        currency = request.form.get('currency')
        urgent = request.form.get('urgent')
        urgentdate = request.form.get('urgentdate')
        service = request.form.get('service')
        serviceOther = request.form.get('ServiceOther')
        itemname = request.form.get('itemname')
        itemOther = request.form.get('itemOther')
        publishedDate = request.form.get('publishedDate')
        clientname = request.form.get('clientname')
        details = request.form.get('details')
        teamsOthers = request.form.get('teamsOthers')

        f_reason = request.form.get('f_reason')
        f_reasonOther = request.form.get('f_reasonOther')
        f_deadline = request.form.get('f_deadline')
        f_chinesename = request.form.get('f_chinesename')
        f_engname = request.form.get('f_engname')
        f_cal = request.form.get('f_cal')
        calOther = request.form.get('calOther')
        f_date = request.form.get('f_date')

        if team == 'ADV':
            team = 'ADV'
        elif team == 'HKET_Adv':
            team = 'HKET Adv'
        elif team == 'iET_A':
            team = 'iET - A'
        elif team == 'iET_B':
            team = 'iET - B'
        elif team == 'iET_C':
            team = 'iET - C'
        elif team == 'iET_D':
            team = 'iET - D'
        elif team == 'iET_E':
            team = 'iET - E'
        elif team == 'PET_A':
            team = 'PET - A'
        elif team == 'PET_B':
            team = 'PET - B'
        elif team == 'PET_C':
            team = 'PET - C'
        elif team == 'PET_D':
            team = 'PET - D'
        elif team == 'PET_E':
            team = 'PET - E'
        elif team == 'SkyPost_Adv':
            team = 'Sky Post Adv'
        elif team == 'Sky_Marketing':
            team = 'Sky Marketing'
        elif team == 'Sky_A':
            team = 'Sky A'
        elif team == 'Sky_B':
            team = 'Sky B'
        elif team == 'Sky_C':
            team = 'Sky C'
        elif team == 'Sky_D':
            team = 'Sky D'
        elif team == 'Sky_E':
            team = 'Sky E'
        elif team == 'UL':
            team = 'UL'
        elif team == 'UM':
            team = 'UM'
        elif team == 'UT':
            team = 'UT'
        elif team == 'UFB':
            team = 'UFB'
        elif team == 'UHK':
            team = 'UHK'
        elif team == 'Others':
            team = teamsOthers
        elif team == 'Ezone':
            team = 'Ezone'
        elif team == 'iMoney':
            team = 'iMoney'
        else:
            team = team



############### Open Related Google Sheet ###################

    front = loginno.partition(' -')[0]
    try:
        if "- C" in loginno:
            spreadsheet_V = client.open('AC002 Vendor')
            g_sheet = client.open('AC002 Vendor').worksheet(front)

        elif "- F" in loginno:
            spreadsheet_F = client.open('AC002 Freelance')
            g_sheet = client.open('AC002 Freelance').worksheet(front)
        elif "- S" in loginno:
            spreadsheet_S = client.open('AC002 Staff')
            g_sheet = client.open('AC002 Staff').worksheet(front)
        else:
            spreadsheet_V = client.open('AC002 Vendor')
            g_sheet = client.open('AC002 Vendor').worksheet(front)

        cell = g_sheet.find(loginno)
        row = cell.row

        status = g_sheet.cell(row, 19).value
        original_ab = g_sheet.cell(row, 14).value

    ############### Update Data###################

        if status != "Delete" and original_ab == appliedby:
            if team !="":
                g_sheet.update_cell(row, 2, team)
            if payto !="":
                g_sheet.update_cell(row, 4, payto)
            if currency !="":
                g_sheet.update_cell(row, 5, currency)
            if amount !="":
                g_sheet.update_cell(row, 6, amount)
            if  adno !="":
                g_sheet.update_cell(row, 7, adno)
            if  clientname !="":
                g_sheet.update_cell(row, 8, clientname)
            if  service !=None:
                if service == 'others':
                    service_final = serviceOther
                else:
                    service_final = service
                g_sheet.update_cell(row, 9, service_final)
            if  itemname !=None:
                if itemname == 'others':
                    itemname_final = itemOther
                else:
                    itemname_final = itemname
                g_sheet.update_cell(row, 10, itemname_final)
            if  publishedDate !="":
                g_sheet.update_cell(row, 11, publishedDate)
            if  details !="":
                g_sheet.update_cell(row, 12, details)
            if  vendorcode !="":
                g_sheet.update_cell(row, 13, vendorcode)
            g_sheet.update_cell(row, 18, 'Update')
            g_sheet.update_cell(row, 19, today)
            g_sheet.update_cell(row, 20, appliedby)

            fmt = cellFormat(
            backgroundColor=color(0.56,0,0.09),
            textFormat=textFormat(foregroundColor=color(1, 1, 1))
            )

            format_cell_range(g_sheet, str(row) , fmt)

            final = ''
            outsource_final = ''

            THIS_FOLDER = os.path.dirname(os.path.abspath(__file__)) + '/AC002'

            os.chdir(THIS_FOLDER)

            for name in glob.glob('*' + loginno + '*'):
                final = name

            revised_ac002 = THIS_FOLDER + '/' + final


            ############### Excel Write data ###################

            book = openpyxl.load_workbook(revised_ac002)

            sheet = book.active

            excel_client_name = sheet['B21'].value
            excel_service = sheet['B22'].value
            excel_itemname = sheet['B23'].value
            excel_publishedDate = sheet['B24'].value
            excel_details = sheet['B25'].value


            if team !="":
                sheet['D8'] = team

            if payto !="":
                sheet['B14'] = payto

            if amount !="":
                amountFinal = "{:,.2f}".format(float(amount))


            if vendorcode !="":
                sheet['B16'] = vendorcode

            if amount !="":
               sheet['C18'] = currency + str(amountFinal)

            if adno !="":
                sheet['D34'] = 'This Amount has been deducted from advertising revenue (Ad #' + adno +')'

            if clientname !="":
                sheet['B21'] = clientname

            if service !=None:
                if service == 'others':
                    sheet['B22'] = serviceOther
                else:
                    sheet['B22'] = service

            if itemname !=None:
                if itemname == 'others':
                    sheet['B23'] = itemOther
                else:
                    sheet['B23'] = itemname

            if publishedDate !="":
                sheet['B24'] = publishedDate

            if details !="":
                sheet['B25'] = details

            if appliedby !="":
                sheet['A34'] = appliedby

            if extno !="":
               sheet['A35'] = 'Applicant\'s Name (Ext. No.' + extno + ' )'

            if payment !="":
                if payment == 'cheque':
                    sheet['A11'] = 'Payment Method :      ☑ Cheque           ☐ T/T           ☐  Other (Specify)'
                elif payment == 'TT':
                    sheet['A11'] = 'Payment Method :      ☐ Cheque           ☑ T/T           ☐  Other (Specify)'
                elif payment == 'other':
                    sheet['A11'] = 'Payment Method :      ☐ Cheque           ☐ T/T           ☑  Other (Specify):'+ otherSpecify
                else:
                    sheet['A11'] = 'Payment Method :      ☑ Cheque           ☐ T/T           ☐  Other (Specify)'

            if mail !="":
                if mail == 'fin' and  urgent == 'yes':
                    sheet['A12'] ='Mail / Send by :       ☑ FIN                     ☐ Applicant             ☑ Urgent (Date Required) :' + urgentdate
                elif mail == 'fin' and  urgent == 'no':
                    sheet['A12'] ='Mail / Send by :       ☑ FIN                     ☐ Applicant             ☐ Urgent (Date Required) :'
                elif mail == 'applicant' and  urgent == 'yes':
                    sheet['A12'] = 'Mail / Send by :       ☐ FIN                     ☑ Applicant             ☑ Urgent (Date Required) :' + urgentdate
                elif mail == 'applicant' and  urgent == 'no':
                    sheet['A12'] = 'Mail / Send by :       ☐ FIN                     ☑ Applicant             ☐ Urgent (Date Required) :'
                else:
                    sheet['A12'] = 'Mail / Send by :       ☐ FIN                     ☑ Applicant             ☐ Urgent (Date Required) '


            medium = Side(border_style="medium", color="000000")
            mediumRed = Side(border_style="medium", color="FF0000")


            sheet['B21'].border = Border(bottom=medium)
            sheet['C21'].border = Border(bottom=medium)
            sheet['D21'].border = Border(bottom=medium)


            sheet['B22'].border = Border(bottom=medium)
            sheet['C22'].border = Border(bottom=medium)
            sheet['D22'].border = Border(bottom=medium)


            sheet['B23'].border = Border(bottom=medium)
            sheet['C23'].border = Border(bottom=medium)
            sheet['D23'].border = Border(bottom=medium)


            sheet['B24'].border = Border(bottom=medium)
            sheet['C24'].border = Border(bottom=medium)
            sheet['D24'].border = Border(bottom=medium)

            sheet['B25'].border = Border(bottom=medium)
            sheet['C25'].border = Border(bottom=medium)
            sheet['D25'].border = Border(bottom=medium)


            sheet['D36'].border = Border(left=mediumRed, right=mediumRed)
            sheet['D37'].border = Border(bottom=mediumRed, left=mediumRed, right=mediumRed)


            book.save(revised_ac002)

            revised_link = 'http://pamelatsui.pythonanywhere.com/AC002/' + final

        ############### OutSource Excel Write data ###################
            outsource_final = 0
            revised_outsource_link = ''



            if "- F" in loginno:
                OutSource_FOLDER = os.path.dirname(os.path.abspath(__file__)) + '/OutSource'

                os.chdir(OutSource_FOLDER)

                for name in glob.glob('*' + loginno + '*'):
                    outsource_final = name

                revised_outsource = OutSource_FOLDER + '/' + outsource_final

                outsource_book = openpyxl.load_workbook(revised_outsource)

                outsource_sheet = outsource_book.active

                if appliedby != "":
                    outsource_sheet['H4'] = appliedby
                if f_reason != "":
                    outsource_sheet['B7'] = f_reason


                if service !=None:
                    if service == 'others':
                        service_final = serviceOther
                    else:
                        service_final = service
                else:
                    service_final = excel_service

                if itemname !=None:
                    if itemname == 'others':
                        itemname_final = itemOther
                    else:
                        itemname_final = itemname
                else:
                    itemname_final = excel_itemname

                if clientname != "":
                    clientname_final = clientname
                else:
                    clientname_final = excel_client_name

                if publishedDate !="":
                    publishedDate_final = publishedDate
                else:
                    publishedDate_final = excel_publishedDate

                if details !="":
                    details_final = details
                else:
                    details_final = excel_details


                if clientname != "" or service != None or itemname !=None or publishedDate !="":
                  outsource_sheet['B14'] = '客戶名稱:' + clientname_final + ',' + '服務性質:' +  service_final + ',' + '項目名稱:' + itemname_final + ',' + '刊登日期:' + publishedDate_final


                if f_deadline != "":
                    outsource_sheet['B17'] = f_deadline
                if amount != "":
                    outsource_sheet['B20'] = currency + str(amountFinal)
                if f_chinesename != "":
                    outsource_sheet['C32'] = f_chinesename
                if f_engname != "":
                   outsource_sheet['H32'] = f_engname
                if f_date != "":
                    outsource_sheet['B34'] = f_date

                if f_cal != "":
                    if f_cal == 'hour':
                        outsource_sheet['D35'] = '按時間'
                    elif f_cal == 'job':
                        outsource_sheet['D35'] = '按工作'
                    else:
                        outsource_sheet['D35'] = calOther

                if amount != "":
                   outsource_sheet['B37'] = currency + str(amountFinal)


                outsource_book.save(revised_outsource)

                revised_outsource_link = 'http://pamelatsui.pythonanywhere.com/OutSource/' + outsource_final
            msg = 'Your request has been processed successfully.'
        elif status != "Delete" and original_ab != appliedby:
            msg = 'Sorry! The orginial applicant is not same as your input.'
        else:
            msg = 'Sorry! Your Record has been deleted.'
    except:
        msg = 'Sorry! Cannot find this reocrd. Please check again.'

    return render_template("changes_result.html", revised_link=revised_link, revised_outsource_link = revised_outsource_link, msg=msg)


@app.route('/delete/')
def delete():
    return render_template("delete.html")


@app.route('/delete_result', methods=['post','get'])
def delete_result():
    msg = ''
    month = (datetime.now() + timedelta(hours=8)).strftime("%b")
    year = (datetime.now() + timedelta(hours=8)).strftime("%y")
    today = (datetime.now() + timedelta(hours=8)).strftime("%d/%m/%y")

    if request.method == 'POST':
        loginno = request.form.get('loginno')
        appliedby = request.form.get('appliedby')

############### Open Related Google Sheet ###################

    front = loginno.partition(' -')[0]
    try:
        if "- C" in loginno:
            spreadsheet_V = client.open('AC002 Vendor')
            g_sheet = client.open('AC002 Vendor').worksheet(front)

        elif "- F" in loginno:
            spreadsheet_F = client.open('AC002 Freelance')
            g_sheet = client.open('AC002 Freelance').worksheet(front)
        elif "- S" in loginno:
            spreadsheet_S = client.open('AC002 Staff')
            g_sheet = client.open('AC002 Staff').worksheet(front)
        else:
            spreadsheet_V = client.open('AC002 Vendor')
            g_sheet = client.open('AC002 Vendor').worksheet(front)


        cell = g_sheet.find(loginno)
        row = cell.row

        original_ab = g_sheet.cell(row, 14).value

    ############### Update Data###################

        if original_ab == appliedby:
            g_sheet.update_cell(row, 18, 'Delete')
            g_sheet.update_cell(row, 19, today)
            g_sheet.update_cell(row, 20, appliedby)

            fmt = cellFormat(
            backgroundColor=color(0.1, 0.1, 0.1),
            textFormat=textFormat(foregroundColor=color(1, 1, 1))
            )
            format_cell_range(g_sheet, str(row) , fmt)

            msg = 'Your request has been processed successfully.'
        else:
            msg = 'Sorry! The orginial applicant is not same as your input.'
    except:
        msg = 'Sorry! Cannot find this record. Please check again.'


    return render_template("delete_result.html", msg=msg)


@app.route('/enquiry/')
def enquiry():
    return render_template("enquiry.html")


@app.route('/enquiry_result', methods=['post','get'])
def enquiry_result():
    msg_S = ''
    msg_F = ''
    msg_V = ''
    values_list_V = []
    values_list_F = []
    values_list_S = []
    df_V = []
    df_F = []
    df_S = []
    vCount = 0
    fCount = 0
    sCount = 0

    month = (datetime.now() + timedelta(hours=8)).strftime("%b")
    year = (datetime.now() + timedelta(hours=8)).strftime("%y")
    today = (datetime.now() + timedelta(hours=8)).strftime("%d/%m/%y")

    if request.method == 'POST':
        appliedby = request.form.get('appliedby')
        password = request.form.get('password')


    worksheet_V_list = ['Jan 21', 'Feb 21', 'Mar 21', 'Apr 21', 'May 21', 'Jun 21', 'Jul 21', 'Aug 21', 'Sep 21', 'Oct 21', 'Nov 21', 'Dec 21']
    worksheet_F_list = ['Jan 21', 'Feb 21', 'Mar 21', 'Apr 21', 'May 21', 'Jun 21', 'Jul 21', 'Aug 21', 'Sep 21', 'Oct 21', 'Nov 21', 'Dec 21']
    worksheet_S_list = ['Jan 21', 'Feb 21', 'Mar 21', 'Apr 21', 'May 21', 'Jun 21', 'Jul 21', 'Aug 21', 'Sep 21', 'Oct 21', 'Nov 21', 'Dec 21']

    try:
        for sheet in worksheet_V_list:
            spreadsheet_V = client.open('AC002 Vendor').worksheet(sheet)
            cell_V = spreadsheet_V.findall(appliedby,in_column=14)

            list_V = []

            if  vCount == 0:
                list_V = [1]
            else:
                list_V = list_V

            vCount = vCount + 1

            for x in cell_V:
                list_V.append(x.row)

            for y in list_V:
                values_list_V.append(spreadsheet_V.row_values(y))


        values_list_V = np.array(values_list_V)
        df_V = pd.DataFrame(data=values_list_V)
        df_V = df_V.loc[df_V['Applied By'] == appliedby]



    except:
        values_list_V = np.array(values_list_V)
        df_V = pd.DataFrame(data=values_list_V)



    try:
        for sheet in worksheet_F_list:
            spreadsheet_F = client.open('AC002 Freelance').worksheet(sheet)
            cell_F = spreadsheet_F.findall(appliedby,in_column=14)
            list_F = []

            if  fCount == 0:
                list_F = [1]
            else:
                list_F = list_F

            fCount = fCount + 1

            for a in cell_F:
                list_F.append(a.row)

            for b in list_F:
                values_list_F.append(spreadsheet_F.row_values(b))

        values_list_F = np.array(values_list_F)
        df_F = pd.DataFrame(data=values_list_F)
        df_F = df_F.loc[df_F['Applied By'] == appliedby]

    except:
        values_list_F = np.array(values_list_F)
        df_F = pd.DataFrame(data=values_list_F)

    try:
        for sheet in worksheet_S_list:
            spreadsheet_S = client.open('AC002 Staff').worksheet(sheet)
            cell_S = spreadsheet_S.findall(appliedby,in_column=14)
            list_S = []

            if  sCount == 0:
                list_S = [1]
            else:
                list_S = list_S

            sCount = sCount + 1

            for c in cell_S:
                list_S.append(c.row)

            for d in list_S:
                values_list_S.append(spreadsheet_S.row_values(d))

        values_list_S = np.array(values_list_S)
        df_S = pd.DataFrame(data=values_list_S)
        df_S = df_S.loc[df_S['Applied By'] == appliedby]


    except:
        values_list_S = np.array(values_list_S)
        df_S = pd.DataFrame(data=values_list_S)


    return render_template("enquiry_result.html", V=[df_V.to_html(classes='data', header=False, index=False)],  F=[df_F.to_html(classes='data', header=False, index=False)],  S=[df_S.to_html(classes='data', header=False, index=False)], msg_F = msg_F, msg_S = msg_S, msg_V = msg_V)

